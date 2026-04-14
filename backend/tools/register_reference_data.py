"""Reference data registration utility.

指定されたExcelから以下を抽出し、data配下に登録する。
- 拠点マスタ（NCR） -> data/site_master_ncr.json
- 依頼シートテンプレ -> data/request_sheet_template.json
"""

from __future__ import annotations

import argparse
import json
import logging
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl


ROOT_DIR = Path(__file__).resolve().parents[2]
DATA_DIR = ROOT_DIR / "data"
SITE_MASTER_PATH = DATA_DIR / "site_master_ncr.json"
REQUEST_TEMPLATE_PATH = DATA_DIR / "request_sheet_template.json"
logger = logging.getLogger(__name__)


def _normalize(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def _to_text(value: Any) -> str | None:
    value = _normalize(value)
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def extract_site_master(workbook: openpyxl.Workbook) -> dict[str, Any]:
    ws = workbook["拠点マスタ（NCR）"]
    headers = [_to_text(c.value) for c in ws[1]]

    records: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        values = [_normalize(c.value) for c in row]
        if not any(v not in (None, "") for v in values):
            continue
        rec: dict[str, Any] = {}
        for header, value in zip(headers, values):
            if not header:
                continue
            rec[header] = value
        records.append(rec)

    return {
        "source_sheet": "拠点マスタ（NCR）",
        "headers": [h for h in headers if h],
        "record_count": len(records),
        "records": records,
    }


def _extract_field_pairs(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    start_row: int,
    end_row: int,
) -> list[dict[str, str]]:
    pairs: list[dict[str, str]] = []
    for row in range(start_row, end_row + 1):
        coordination = _to_text(ws.cell(row=row, column=4).value)
        onsite = _to_text(ws.cell(row=row, column=6).value)
        if not coordination and not onsite:
            continue
        pairs.append({
            "coordination": coordination or "",
            "onsite": onsite or "",
        })
    return pairs


def _extract_notes(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    rows: list[int],
    column: int = 4,
) -> list[str]:
    notes: list[str] = []
    for row in rows:
        text = _to_text(ws.cell(row=row, column=column).value)
        if text:
            notes.append(text)
    return notes


def extract_request_sheet_template(workbook: openpyxl.Workbook) -> dict[str, Any]:
    ws = workbook["依頼シート"]

    survey_fields = _extract_field_pairs(ws, 53, 63)
    install_fields = _extract_field_pairs(ws, 74, 84)

    template = {
        "title": _to_text(ws["A2"].value) or "ブリーフィングシート",
        "sections": [
            {
                "id": "common",
                "name": "いつも同じ情報のページ",
                "notes": _extract_notes(ws, [4, 25], column=4),
            },
            {
                "id": "survey",
                "name": "現調",
                "usage_notes": _extract_notes(ws, [48, 49, 50], column=4),
                "fields": survey_fields,
            },
            {
                "id": "install",
                "name": "設置",
                "usage_notes": _extract_notes(ws, [68, 69, 70, 71], column=4),
                "fields": install_fields,
            },
            {
                "id": "equipment_info",
                "name": "新規納入機材情報",
                "notes": _extract_notes(ws, [89, 90, 91], column=4),
            },
            {
                "id": "report_form_link",
                "name": "Formへのリンク（速報入力用）",
                "notes": _extract_notes(ws, [110, 111, 112], column=4),
            },
            {
                "id": "overflow_note",
                "name": "連携事項が多すぎて収まらないとき",
                "notes": _extract_notes(ws, [130], column=1),
            },
        ],
    }

    return {
        "source_sheet": "依頼シート",
        "template": template,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Register site master and request sheet template from Excel.")
    parser.add_argument("excel_path", help="source xlsx path")
    args = parser.parse_args()

    source_path = Path(args.excel_path).expanduser().resolve()
    if not source_path.exists():
        raise FileNotFoundError(f"Excel file not found: {source_path}")

    wb = openpyxl.load_workbook(source_path, data_only=True)
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    now = datetime.now().isoformat()

    site_payload = {
        "source_file": source_path.name,
        "imported_at": now,
        **extract_site_master(wb),
    }
    SITE_MASTER_PATH.write_text(
        json.dumps(site_payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    request_payload = {
        "source_file": source_path.name,
        "imported_at": now,
        **extract_request_sheet_template(wb),
    }
    REQUEST_TEMPLATE_PATH.write_text(
        json.dumps(request_payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    logger.info("Registered: %s", SITE_MASTER_PATH)
    logger.info("Registered: %s", REQUEST_TEMPLATE_PATH)


if __name__ == "__main__":
    main()
