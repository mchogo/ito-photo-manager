"""Excel報告書の自動生成

openpyxlを使用して、案件データと写真を埋め込んだExcelファイルを生成する。
"""

from __future__ import annotations

import io
import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from PIL import Image as PILImage

logger = logging.getLogger(__name__)

# Excel挿入用画像の最大幅（ピクセル）
EXCEL_IMAGE_MAX_WIDTH = 300
# 画像を配置するセルの高さ（ポイント）
IMAGE_ROW_HEIGHT = 180
# 画像を配置するセルの幅（文字数）
IMAGE_COL_WIDTH = 45

# スタイル定義
HEADER_FONT = Font(name="メイリオ", bold=True, size=14)
SUBHEADER_FONT = Font(name="メイリオ", bold=True, size=11)
LABEL_FONT = Font(name="メイリオ", size=10)
VALUE_FONT = Font(name="メイリオ", size=10)
EQ_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
EQ_FONT = Font(name="メイリオ", bold=True, size=11, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _resize_for_excel(photo_path: Path) -> bytes:
    """Excel挿入用に画像をリサイズする"""
    img = PILImage.open(photo_path)
    if img.width > EXCEL_IMAGE_MAX_WIDTH:
        ratio = EXCEL_IMAGE_MAX_WIDTH / img.width
        new_height = int(img.height * ratio)
        img = img.resize((EXCEL_IMAGE_MAX_WIDTH, new_height), PILImage.Resampling.LANCZOS)
    if img.mode in ("RGBA", "P"):
        img = img.convert("RGB")
    buf = io.BytesIO()
    img.save(buf, format="JPEG", quality=80, optimize=True)
    buf.seek(0)
    return buf


def generate_excel(project: dict, photos_dir: Path) -> bytes:
    """案件データからExcelファイルを生成し、バイト列で返す

    レイアウト:
    - 行1: タイトル
    - 行2: 現場ID
    - 行3: 作業日
    - 行4: 作業員名
    - 行5: 空行
    - 行6〜: 機器セクション（機器名ヘッダ → 各スロット: ラベル + 画像）
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "撮影報告書"

    # 列幅設定
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = IMAGE_COL_WIDTH

    # --- ヘッダ情報 ---
    row = 1
    ws.cell(row=row, column=1, value="現場撮影報告書").font = HEADER_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)

    row = 2
    ws.cell(row=row, column=1, value="現場ID").font = LABEL_FONT
    ws.cell(row=row, column=1).border = THIN_BORDER
    ws.cell(row=row, column=2, value=project["site_id"]).font = VALUE_FONT
    ws.cell(row=row, column=2).border = THIN_BORDER

    row = 3
    ws.cell(row=row, column=1, value="作業日").font = LABEL_FONT
    ws.cell(row=row, column=1).border = THIN_BORDER
    ws.cell(row=row, column=2, value=project["work_date"]).font = VALUE_FONT
    ws.cell(row=row, column=2).border = THIN_BORDER

    row = 4
    ws.cell(row=row, column=1, value="作業員名").font = LABEL_FONT
    ws.cell(row=row, column=1).border = THIN_BORDER
    ws.cell(row=row, column=2, value=project["worker_name"]).font = VALUE_FONT
    ws.cell(row=row, column=2).border = THIN_BORDER

    row = 6  # 空行を1行あけて機器セクション開始

    # --- 機器セクション ---
    for eq in project["equipment"]:
        # 機器名ヘッダ
        cell_a = ws.cell(row=row, column=1, value=eq["name"])
        cell_a.font = EQ_FONT
        cell_a.fill = EQ_FILL
        cell_a.border = THIN_BORDER
        cell_a.alignment = Alignment(vertical="center")
        cell_b = ws.cell(row=row, column=2)
        cell_b.fill = EQ_FILL
        cell_b.border = THIN_BORDER
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        row += 1

        for slot in eq["slots"]:
            # スロットラベル
            label_cell = ws.cell(row=row, column=1, value=slot["label"])
            label_cell.font = LABEL_FONT
            label_cell.border = THIN_BORDER
            label_cell.alignment = Alignment(vertical="top")

            photo_cell = ws.cell(row=row, column=2)
            photo_cell.border = THIN_BORDER

            # 写真挿入
            if slot["photo_filename"]:
                photo_path = photos_dir / slot["photo_filename"]
                if photo_path.exists():
                    try:
                        img_buf = _resize_for_excel(photo_path)
                        xl_img = XlImage(img_buf)
                        # セル位置に画像を配置
                        anchor = f"B{row}"
                        ws.add_image(xl_img, anchor)
                        ws.row_dimensions[row].height = IMAGE_ROW_HEIGHT
                    except Exception:
                        logger.exception("Failed to embed image: %s", photo_path)
                        photo_cell.value = f"[画像読込エラー: {slot['photo_filename']}]"
                else:
                    photo_cell.value = f"[ファイル未検出: {slot['photo_filename']}]"
            else:
                photo_cell.value = "[未撮影]"
                photo_cell.font = Font(name="メイリオ", size=10, color="FF0000")

            row += 1

        row += 1  # 機器間に空行

    # バイト列として出力
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
