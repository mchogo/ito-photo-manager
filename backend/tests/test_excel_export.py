"""Excel出力のテスト"""

import io
from datetime import date
from pathlib import Path
from unittest import mock

import pytest
from openpyxl import load_workbook
from PIL import Image

import storage
from excel_export import generate_excel


@pytest.fixture(autouse=True)
def temp_data_dir(tmp_path):
    """テスト用の一時データディレクトリ"""
    projects_dir = tmp_path / "projects"
    photos_dir = tmp_path / "photos"
    projects_dir.mkdir()
    photos_dir.mkdir()
    with mock.patch.object(storage, "PROJECTS_DIR", projects_dir), \
         mock.patch.object(storage, "PHOTOS_DIR", photos_dir):
        yield tmp_path


def _create_dummy_jpeg(path: Path) -> None:
    """テスト用JPEG画像を生成"""
    img = Image.new("RGB", (400, 300), color="blue")
    img.save(str(path), format="JPEG")


def test_excel_with_no_photos():
    """写真なしでもExcelが生成できる"""
    project = storage.create_project(
        site_id="SITE-X",
        work_date=date(2025, 3, 1),
        worker_name="山田花子",
        equipment_ids=["pos_register"],
    )
    result = generate_excel(project, storage.PHOTOS_DIR)
    assert len(result) > 0
    # openpyxlで読み込めることを確認
    wb = load_workbook(io.BytesIO(result))
    ws = wb.active
    assert ws.title == "撮影報告書"
    assert ws["A1"].value == "現場撮影報告書"
    assert ws["B2"].value == "SITE-X"
    assert ws["B3"].value == "2025-03-01"
    assert ws["B4"].value == "山田花子"


def test_excel_with_photos():
    """写真ありでExcelが生成できる"""
    project = storage.create_project(
        site_id="SITE-Y",
        work_date=date(2025, 4, 10),
        worker_name="鈴木一郎",
        equipment_ids=["router"],
    )
    # ダミー写真を設置
    pid = project["project_id"]
    img = Image.new("RGB", (400, 300), color="green")
    buf = io.BytesIO()
    img.save(buf, format="JPEG")
    storage.save_photo(
        pid, "router", "router_front", buf.getvalue(), "test.jpg",
    )

    updated = storage.get_project(pid)
    result = generate_excel(updated, storage.PHOTOS_DIR)
    assert len(result) > 0
    wb = load_workbook(io.BytesIO(result))
    ws = wb.active
    # 画像が挿入されていることを確認（openpyxlのimages属性）
    assert len(ws._images) >= 1


def test_excel_multiple_equipment():
    """複数機器でExcelが正しく生成できる"""
    project = storage.create_project(
        site_id="SITE-Z",
        work_date=date(2025, 5, 20),
        worker_name="佐藤次郎",
        equipment_ids=["pos_register", "cash_drawer", "router"],
    )
    result = generate_excel(project, storage.PHOTOS_DIR)
    wb = load_workbook(io.BytesIO(result))
    ws = wb.active
    # 全セルの値にPOS、キャッシュドロア、ルーターが含まれること
    values = [cell.value for row in ws.iter_rows() for cell in row if cell.value]
    assert "POSレジ本体" in values
    assert "キャッシュドロア" in values
    assert "ルーター" in values
